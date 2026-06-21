from typing import Any, Dict
import openpyxl


def parse_xlsx(file_path: str) -> Dict[str, Any]:
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheets = {}
    text_parts = []

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [str(h) if h is not None else "" for h in rows[0]]
        data_rows = []
        for row in rows[1:]:
            row_dict = {headers[i]: (str(v) if v is not None else "") for i, v in enumerate(row)}
            data_rows.append(row_dict)

        sheets[sheet_name] = {
            "columns": headers,
            "row_count": len(data_rows),
            "rows": data_rows,
        }

        text_parts.append(f"Sheet: {sheet_name}")
        for i, row in enumerate(data_rows, start=1):
            row_text = ", ".join(f"{k}: {v}" for k, v in row.items())
            text_parts.append(f"Row {i}: {row_text}")

    return {
        "content_type": "spreadsheet",
        "sheet_count": len(sheets),
        "sheets": sheets,
        "text": "\n".join(text_parts),
    }